from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.db import (
    CategoriaFinanceira,
    CategoriaProduto,
    Cupom,
    FormaPagamento,
    Funcionario,
    FuncionarioEmpresa,
    Produto,
    ProdutoEmpresa,
    TipoCategoriaFinanceira,
    TipoDesconto,
)
from app.repositorys.import_export_repository import ImportExportRepository
from app.security.password import hash_password
from app.services.acesso_empresa_service import AcessoEmpresaService
from app.services.time_service import TimeService


class ImportExportService:
    # Mapeamento central dos layouts para manter importacao, exportacao e template em sincronia.
    ENTITY_DEFINITIONS = {
        "categorias": {
            "label": "Categorias de produtos",
            "description": "Cadastro base para organizar o catalogo de estoque e PDV.",
            "icon": "layers",
            "supports_company_filter": False,
            "import_permission": "criar_categoria",
            "export_permission": "visualizar_categoria",
            "columns": [
                {"key": "nome", "header": "nome", "required": True, "example": "Bebidas"},
                {"key": "descricao", "header": "descricao", "required": False, "example": "Produtos gelados"},
                {"key": "ativo", "header": "ativo", "required": False, "example": "SIM"},
            ],
        },
        "produtos": {
            "label": "Produtos",
            "description": "Cadastro em lote com empresa, categoria, estoque minimo e valores.",
            "icon": "package",
            "supports_company_filter": True,
            "import_permission": "criar_produto",
            "export_permission": "visualizar_produto",
            "columns": [
                {"key": "empresa", "header": "empresa", "required": True, "example": "BlueOcean Centro"},
                {"key": "categoria", "header": "categoria", "required": False, "example": "Bebidas"},
                {"key": "nome", "header": "nome", "required": True, "example": "Refrigerante 2L"},
                {"key": "descricao", "header": "descricao", "required": False, "example": "Sabor cola"},
                {"key": "codigo_barras", "header": "codigo_barras", "required": False, "example": "789000000001"},
                {"key": "possui_ncm", "header": "possui_ncm", "required": False, "example": "NAO"},
                {"key": "ncm", "header": "ncm", "required": False, "example": "2202.10.00"},
                {"key": "estoque_atual", "header": "estoque_atual", "required": False, "example": "30"},
                {"key": "estoque_minimo", "header": "estoque_minimo", "required": False, "example": "5"},
                {"key": "valor_compra", "header": "valor_compra", "required": False, "example": "4,50"},
                {"key": "valor_venda", "header": "valor_venda", "required": False, "example": "7,50"},
                {"key": "data_validade", "header": "data_validade", "required": False, "example": "2026-12-31"},
                {"key": "ativo", "header": "ativo", "required": False, "example": "SIM"},
            ],
        },
        "funcionarios": {
            "label": "Funcionarios",
            "description": "Importe equipe, vinculos por empresa, salario e meta comercial.",
            "icon": "users",
            "supports_company_filter": True,
            "import_permission": "criar_funcionario",
            "export_permission": "visualizar_funcionario",
            "columns": [
                {"key": "empresa", "header": "empresa", "required": True, "example": "BlueOcean Centro"},
                {"key": "role", "header": "role", "required": True, "example": "operador"},
                {"key": "nome", "header": "nome", "required": True, "example": "Maria Silva"},
                {"key": "cpf", "header": "cpf", "required": True, "example": "123.456.789-10"},
                {"key": "usuario", "header": "usuario", "required": True, "example": "maria.silva"},
                {"key": "senha", "header": "senha", "required": False, "example": "123456"},
                {"key": "salario", "header": "salario", "required": False, "example": "2500,00"},
                {"key": "meta", "header": "meta", "required": False, "example": "18000,00"},
                {"key": "ativo", "header": "ativo", "required": False, "example": "SIM"},
            ],
        },
        "cupons": {
            "label": "Cupons",
            "description": "Campanhas em lote com validade e desconto por percentual ou valor.",
            "icon": "ticket-percent",
            "supports_company_filter": False,
            "import_permission": "criar_cupom",
            "export_permission": "visualizar_cupom",
            "columns": [
                {"key": "nome", "header": "nome", "required": True, "example": "Inauguracao"},
                {"key": "codigo", "header": "codigo", "required": True, "example": "INAUGURA10"},
                {"key": "data_validade", "header": "data_validade", "required": True, "example": "2026-12-31"},
                {"key": "tipo_desconto", "header": "tipo_desconto", "required": True, "example": "PERCENTUAL"},
                {"key": "valor_desconto", "header": "valor_desconto", "required": True, "example": "10"},
                {"key": "ativo", "header": "ativo", "required": False, "example": "SIM"},
            ],
        },
        "formas_pagamento": {
            "label": "Formas de pagamento",
            "description": "Cadastros auxiliares usados no PDV, financeiro e vales.",
            "icon": "wallet-cards",
            "supports_company_filter": False,
            "import_permission": "criar_lancamento_financeiro",
            "export_permission": "visualizar_financeiro",
            "columns": [
                {"key": "nome", "header": "nome", "required": True, "example": "Pix"},
                {"key": "ativo", "header": "ativo", "required": False, "example": "SIM"},
            ],
        },
        "categorias_financeiras": {
            "label": "Categorias financeiras",
            "description": "Entradas e saidas usadas nos lancamentos e no controle de caixa.",
            "icon": "landmark",
            "supports_company_filter": False,
            "import_permission": "criar_lancamento_financeiro",
            "export_permission": "visualizar_financeiro",
            "columns": [
                {"key": "tipo_categoria", "header": "tipo_categoria", "required": True, "example": "SAIDA"},
                {"key": "nome", "header": "nome", "required": True, "example": "Despesas operacionais"},
                {"key": "ativo", "header": "ativo", "required": False, "example": "SIM"},
            ],
        },
    }

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="0EA5E9")
    SECTION_FILL = PatternFill(fill_type="solid", fgColor="0F172A")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    TITLE_FONT = Font(color="0F172A", bold=True, size=14)
    GENERIC_OPERATION_PERMISSIONS = {
        "import": "importar_dados_cadastrais",
        "export": "exportar_dados_cadastrais",
    }
    MAX_IMPORT_ROWS = 5000

    @classmethod
    def obter_contexto_painel(cls, tenant_id, escopo):
        empresas = ImportExportRepository.listar_empresas(
            tenant_id,
            AcessoEmpresaService.filtrar_empresa_ids(escopo),
        )
        entidades = []

        for codigo, config in cls.ENTITY_DEFINITIONS.items():
            entidades.append({
                "codigo": codigo,
                "nome": config["label"],
                "descricao": config["description"],
                "icone": config["icon"],
                "colunas": [
                    {
                        "codigo": coluna["key"],
                        "nome": coluna["header"],
                        "obrigatoria": coluna["required"],
                    }
                    for coluna in config["columns"]
                ],
                "filtra_empresa": config["supports_company_filter"],
                "pode_importar": cls._pode_operar_entidade(codigo, escopo, "import"),
                "pode_exportar": cls._pode_operar_entidade(codigo, escopo, "export"),
            })

        return {
            "entidades": entidades,
            "empresas": [
                {
                    "id": item.id,
                    "nome": item.nome_fantasia,
                    "cnpj": item.cnpj,
                }
                for item in empresas
            ],
            "resumo": {
                "entidades": len(entidades),
                "empresas": len(empresas),
                "modelos": len(entidades),
            },
        }

    @classmethod
    def gerar_template(cls, entidade, tenant_id, escopo):
        config = cls._get_entity_config(entidade)
        cls._validate_operation(entidade, escopo, "import")

        workbook = Workbook()
        dados = workbook.active
        dados.title = "dados"
        headers = [coluna["header"] for coluna in config["columns"]]
        dados.append(headers)
        cls._style_header_row(dados, 1)
        dados.freeze_panes = "A2"

        instrucoes = workbook.create_sheet("instrucoes")
        cls._fill_instructions_sheet(instrucoes, config)

        referencias = workbook.create_sheet("referencias")
        cls._fill_reference_sheet(referencias, entidade, tenant_id, escopo)

        cls._auto_fit_sheet(dados)
        cls._auto_fit_sheet(instrucoes)
        cls._auto_fit_sheet(referencias)

        return {
            "filename": f"modelo_importacao_{entidade}_{date.today().isoformat()}.xlsx",
            "content": cls._workbook_to_bytes(workbook),
        }

    @classmethod
    def exportar_entidade(cls, entidade, tenant_id, escopo, empresa_id=None):
        config = cls._get_entity_config(entidade)
        cls._validate_operation(entidade, escopo, "export")

        if config["supports_company_filter"] and empresa_id:
            AcessoEmpresaService.validar_empresa(empresa_id, escopo)

        rows = cls._build_export_rows(entidade, tenant_id, escopo, empresa_id)

        workbook = Workbook()
        dados = workbook.active
        dados.title = "dados"
        headers = [coluna["header"] for coluna in config["columns"]]
        dados.append(headers)
        cls._style_header_row(dados, 1)
        dados.freeze_panes = "A2"

        for row in rows:
            dados.append([row.get(coluna["key"], "") for coluna in config["columns"]])

        info = workbook.create_sheet("informacoes")
        info["A1"] = "Exportacao gerada automaticamente"
        info["A1"].font = cls.TITLE_FONT
        info["A3"] = "Entidade"
        info["B3"] = config["label"]
        info["A4"] = "Linhas exportadas"
        info["B4"] = len(rows)
        info["A5"] = "Gerado em"
        info["B5"] = TimeService.now_utc_naive().strftime("%Y-%m-%d %H:%M:%S")

        cls._auto_fit_sheet(dados)
        cls._auto_fit_sheet(info)

        return {
            "filename": f"exportacao_{entidade}_{date.today().isoformat()}.xlsx",
            "content": cls._workbook_to_bytes(workbook),
        }

    @classmethod
    def importar_entidade(cls, entidade, arquivo, tenant_id, escopo, funcionario_id):
        cls._get_entity_config(entidade)
        cls._validate_operation(entidade, escopo, "import")

        if not arquivo:
            raise ValueError("Selecione um arquivo XLSX para importar.")

        nome_arquivo = (getattr(arquivo, "filename", "") or "").lower()
        if not nome_arquivo.endswith((".xlsx", ".xlsm")):
            raise ValueError("Formato invalido. Use um arquivo .xlsx.")

        workbook = load_workbook(arquivo, data_only=True)
        sheet = cls._get_data_sheet(workbook)
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValueError("A planilha nao possui dados para importar.")
        if len(rows) - 1 > cls.MAX_IMPORT_ROWS:
            raise ValueError(f"A planilha excede o limite de {cls.MAX_IMPORT_ROWS} linhas por importacao.")

        config = cls._get_entity_config(entidade)
        index_map = cls._build_header_map(rows[0], config["columns"])

        summary = {
            "entidade": entidade,
            "nome_entidade": config["label"],
            "processadas": 0,
            "sucesso": 0,
            "falhas": 0,
            "criadas": 0,
            "atualizadas": 0,
            "erros": [],
        }

        for excel_row_number, values in enumerate(rows[1:], start=2):
            if cls._row_is_empty(values):
                continue

            summary["processadas"] += 1
            row = cls._extract_row(values, index_map)

            try:
                action = cls._dispatch_import(entidade, row, tenant_id, escopo, funcionario_id)
                ImportExportRepository.salvar()
                summary["sucesso"] += 1
                if action == "criado":
                    summary["criadas"] += 1
                else:
                    summary["atualizadas"] += 1
            except Exception as exc:
                ImportExportRepository.rollback()
                summary["falhas"] += 1
                summary["erros"].append({
                    "linha": excel_row_number,
                    "mensagem": str(exc),
                })

        return summary

    @classmethod
    def _dispatch_import(cls, entidade, row, tenant_id, escopo, funcionario_id):
        handlers = {
            "categorias": cls._import_categoria,
            "produtos": cls._import_produto,
            "funcionarios": cls._import_funcionario,
            "cupons": cls._import_cupom,
            "formas_pagamento": cls._import_forma_pagamento,
            "categorias_financeiras": cls._import_categoria_financeira,
        }
        return handlers[entidade](row, tenant_id, escopo, funcionario_id)

    @classmethod
    def _import_categoria(cls, row, tenant_id, escopo, funcionario_id):
        del escopo, funcionario_id

        nome = cls._required_text(row.get("nome"), "nome")
        descricao = cls._optional_text(row.get("descricao"))
        ativo = cls._parse_bool(row.get("ativo"), default=True)

        categoria = ImportExportRepository.buscar_categoria_produto_por_nome(tenant_id, nome)
        if categoria:
            categoria.descricao = descricao
            categoria.ativo = ativo
            categoria.atualizado_em = TimeService.now_utc_naive()
            return "atualizado"

        categoria = CategoriaProduto(
            tenant_id=tenant_id,
            nome=nome,
            descricao=descricao,
            ativo=ativo,
            criado_em=TimeService.now_utc_naive(),
            atualizado_em=TimeService.now_utc_naive(),
        )
        ImportExportRepository.adicionar(categoria)
        return "criado"

    @classmethod
    def _import_produto(cls, row, tenant_id, escopo, funcionario_id):
        empresa = cls._resolve_empresa(tenant_id, escopo, row.get("empresa"))
        nome = cls._required_text(row.get("nome"), "nome")
        codigo_barras = cls._optional_text(row.get("codigo_barras"))
        descricao = cls._optional_text(row.get("descricao"))
        categoria_nome = cls._optional_text(row.get("categoria"))
        possui_ncm = cls._parse_bool(row.get("possui_ncm"), default=False)
        ncm = cls._optional_text(row.get("ncm"))
        estoque_atual = cls._parse_int(row.get("estoque_atual"), default=0, field_name="estoque_atual")
        estoque_minimo = cls._parse_int(row.get("estoque_minimo"), default=0, field_name="estoque_minimo")
        valor_compra = cls._parse_decimal(row.get("valor_compra"), default="0.00", field_name="valor_compra")
        valor_venda = cls._parse_decimal(row.get("valor_venda"), default="0.00", field_name="valor_venda")
        data_validade = cls._parse_optional_date(row.get("data_validade"), "data_validade")
        ativo = cls._parse_bool(row.get("ativo"), default=True)

        if possui_ncm and not ncm:
            raise ValueError("Informe o NCM quando possui_ncm estiver marcado como SIM.")

        categoria = None
        if categoria_nome:
            categoria = ImportExportRepository.buscar_categoria_produto_por_nome(tenant_id, categoria_nome)
            if not categoria:
                categoria = CategoriaProduto(
                    tenant_id=tenant_id,
                    nome=categoria_nome,
                    descricao=None,
                    ativo=True,
                    criado_em=TimeService.now_utc_naive(),
                    atualizado_em=TimeService.now_utc_naive(),
                )
                ImportExportRepository.adicionar(categoria)
                ImportExportRepository.flush()

        produto_por_codigo = ImportExportRepository.buscar_produto_por_codigo_barras(tenant_id, codigo_barras)
        produto_por_nome = ImportExportRepository.buscar_produto_por_nome(tenant_id, nome)

        if produto_por_codigo and produto_por_nome and produto_por_codigo.id != produto_por_nome.id:
            raise ValueError("Ha conflito entre nome e codigo de barras apontando para produtos diferentes.")

        produto = produto_por_codigo or produto_por_nome
        acao = "criado"

        if produto:
            produto.nome = nome
            produto.descricao = descricao
            produto.codigo_barras = codigo_barras
            produto.categoria_id = categoria.id if categoria else None
            produto.possui_ncm = possui_ncm
            produto.ncm = ncm
            produto.ativo = ativo
            produto.atualizado_em = TimeService.now_utc_naive()
            acao = "atualizado"
        else:
            produto = Produto(
                tenant_id=tenant_id,
                categoria_id=categoria.id if categoria else None,
                criado_por_funcionario_id=funcionario_id,
                nome=nome,
                descricao=descricao,
                possui_ncm=possui_ncm,
                ncm=ncm,
                codigo_barras=codigo_barras,
                ativo=ativo,
                criado_em=TimeService.now_utc_naive(),
                atualizado_em=TimeService.now_utc_naive(),
            )
            ImportExportRepository.adicionar(produto)
            ImportExportRepository.flush()

        produto_empresa = ImportExportRepository.buscar_produto_empresa(tenant_id, produto.id, empresa.id)
        if produto_empresa:
            produto_empresa.estoque_atual = estoque_atual
            produto_empresa.estoque_minimo = estoque_minimo
            produto_empresa.valor_compra = valor_compra
            produto_empresa.valor_venda = valor_venda
            produto_empresa.valor_varejo = valor_venda
            produto_empresa.valor_atacado = valor_venda
            produto_empresa.quantidade_minima_atacado = 1
            produto_empresa.data_validade = data_validade
            produto_empresa.ativo = ativo
            produto_empresa.atualizado_em = TimeService.now_utc_naive()
            return "atualizado"

        produto_empresa = ProdutoEmpresa(
            tenant_id=tenant_id,
            produto_id=produto.id,
            empresa_id=empresa.id,
            estoque_atual=estoque_atual,
            estoque_minimo=estoque_minimo,
            valor_compra=valor_compra,
            valor_venda=valor_venda,
            valor_varejo=valor_venda,
            valor_atacado=valor_venda,
            quantidade_minima_atacado=1,
            data_validade=data_validade,
            ativo=ativo,
            criado_em=TimeService.now_utc_naive(),
            atualizado_em=TimeService.now_utc_naive(),
        )
        ImportExportRepository.adicionar(produto_empresa)
        return acao

    @classmethod
    def _import_funcionario(cls, row, tenant_id, escopo, funcionario_id):
        del funcionario_id

        empresa = cls._resolve_empresa(tenant_id, escopo, row.get("empresa"))
        role = cls._resolve_role(tenant_id, row.get("role"))
        nome = cls._required_text(row.get("nome"), "nome")
        cpf = cls._required_text(row.get("cpf"), "cpf")
        usuario = cls._required_text(row.get("usuario"), "usuario")
        senha = cls._optional_text(row.get("senha"))
        salario = cls._parse_decimal(row.get("salario"), default="0.00", field_name="salario")
        meta = cls._parse_decimal(row.get("meta"), default="0.00", field_name="meta")
        ativo = cls._parse_bool(row.get("ativo"), default=True)

        funcionario_por_cpf = ImportExportRepository.buscar_funcionario_por_cpf(tenant_id, cpf)
        funcionario_por_usuario = ImportExportRepository.buscar_funcionario_por_usuario(tenant_id, usuario)

        if funcionario_por_cpf and funcionario_por_usuario and funcionario_por_cpf.id != funcionario_por_usuario.id:
            raise ValueError("CPF e usuario informados pertencem a funcionarios diferentes.")

        funcionario = funcionario_por_cpf or funcionario_por_usuario
        acao = "criado"

        if funcionario:
            funcionario.nome = nome
            funcionario.cpf = cpf
            funcionario.usuario = usuario
            funcionario.role_id = role.id
            funcionario.salario = salario
            funcionario.meta = meta
            funcionario.ativo = ativo
            funcionario.atualizado_em = TimeService.now_utc_naive()
            if senha:
                funcionario.senha_hash = hash_password(senha)
            acao = "atualizado"
        else:
            if not senha:
                raise ValueError("Informe a senha para novos funcionarios.")

            funcionario = Funcionario(
                tenant_id=tenant_id,
                role_id=role.id,
                nome=nome,
                cpf=cpf,
                usuario=usuario,
                senha_hash=hash_password(senha),
                salario=salario,
                meta=meta,
                ativo=ativo,
                criado_em=TimeService.now_utc_naive(),
                atualizado_em=TimeService.now_utc_naive(),
            )
            ImportExportRepository.adicionar(funcionario)
            ImportExportRepository.flush()

        vinculo = ImportExportRepository.buscar_funcionario_empresa(tenant_id, funcionario.id, empresa.id)
        if vinculo:
            vinculo.ativo = ativo
            vinculo.atualizado_em = TimeService.now_utc_naive()
            return "atualizado"

        vinculo = FuncionarioEmpresa(
            tenant_id=tenant_id,
            funcionario_id=funcionario.id,
            empresa_id=empresa.id,
            ativo=ativo,
            criado_em=TimeService.now_utc_naive(),
            atualizado_em=TimeService.now_utc_naive(),
        )
        ImportExportRepository.adicionar(vinculo)
        return acao

    @classmethod
    def _import_cupom(cls, row, tenant_id, escopo, funcionario_id):
        del escopo

        nome = cls._required_text(row.get("nome"), "nome")
        codigo = cls._required_text(row.get("codigo"), "codigo").upper()
        data_validade = cls._parse_required_date(row.get("data_validade"), "data_validade")
        tipo_desconto = cls._parse_tipo_desconto(row.get("tipo_desconto"))
        valor_desconto = cls._parse_decimal(row.get("valor_desconto"), default=None, field_name="valor_desconto")
        ativo = cls._parse_bool(row.get("ativo"), default=True)

        if valor_desconto <= 0:
            raise ValueError("valor_desconto deve ser maior que zero.")
        if tipo_desconto == TipoDesconto.PERCENTUAL and valor_desconto > Decimal("100.00"):
            raise ValueError("valor_desconto percentual nao pode ultrapassar 100.")
        if data_validade < date.today():
            raise ValueError("data_validade nao pode estar no passado.")

        cupom = ImportExportRepository.buscar_cupom_por_codigo(tenant_id, codigo)
        if cupom:
            cupom.nome = nome
            cupom.data_validade = data_validade
            cupom.tipo_desconto = tipo_desconto
            cupom.valor_desconto = valor_desconto
            cupom.ativo = ativo
            cupom.atualizado_em = TimeService.now_utc_naive()
            return "atualizado"

        cupom = Cupom(
            tenant_id=tenant_id,
            criado_por_funcionario_id=funcionario_id,
            nome=nome,
            codigo=codigo,
            data_validade=data_validade,
            tipo_desconto=tipo_desconto,
            valor_desconto=valor_desconto,
            ativo=ativo,
            criado_em=TimeService.now_utc_naive(),
            atualizado_em=TimeService.now_utc_naive(),
        )
        ImportExportRepository.adicionar(cupom)
        return "criado"

    @classmethod
    def _import_forma_pagamento(cls, row, tenant_id, escopo, funcionario_id):
        del escopo, funcionario_id

        nome = cls._required_text(row.get("nome"), "nome")
        ativo = cls._parse_bool(row.get("ativo"), default=True)

        forma = ImportExportRepository.buscar_forma_pagamento_por_nome(tenant_id, nome)
        if forma:
            forma.ativo = ativo
            forma.atualizado_em = TimeService.now_utc_naive()
            return "atualizado"

        forma = FormaPagamento(
            tenant_id=tenant_id,
            nome=nome,
            ativo=ativo,
            criado_em=TimeService.now_utc_naive(),
            atualizado_em=TimeService.now_utc_naive(),
        )
        ImportExportRepository.adicionar(forma)
        return "criado"

    @classmethod
    def _import_categoria_financeira(cls, row, tenant_id, escopo, funcionario_id):
        del escopo, funcionario_id

        nome = cls._required_text(row.get("nome"), "nome")
        tipo_categoria = cls._parse_tipo_categoria(row.get("tipo_categoria"))
        ativo = cls._parse_bool(row.get("ativo"), default=True)

        categoria = ImportExportRepository.buscar_categoria_financeira_por_nome_tipo(
            tenant_id,
            nome,
            tipo_categoria,
        )
        if categoria:
            categoria.ativo = ativo
            categoria.atualizado_em = TimeService.now_utc_naive()
            return "atualizado"

        categoria = CategoriaFinanceira(
            tenant_id=tenant_id,
            nome=nome,
            tipo_categoria=tipo_categoria,
            ativo=ativo,
            criado_em=TimeService.now_utc_naive(),
            atualizado_em=TimeService.now_utc_naive(),
        )
        ImportExportRepository.adicionar(categoria)
        return "criado"

    @classmethod
    def _build_export_rows(cls, entidade, tenant_id, escopo, empresa_id=None):
        builders = {
            "categorias": cls._export_categorias,
            "produtos": cls._export_produtos,
            "funcionarios": cls._export_funcionarios,
            "cupons": cls._export_cupons,
            "formas_pagamento": cls._export_formas_pagamento,
            "categorias_financeiras": cls._export_categorias_financeiras,
        }
        return builders[entidade](tenant_id, escopo, empresa_id)

    @classmethod
    def _export_categorias(cls, tenant_id, escopo, empresa_id=None):
        del escopo, empresa_id
        categorias = ImportExportRepository.listar_categorias_produto(tenant_id)
        return [
            {
                "nome": item.nome,
                "descricao": item.descricao or "",
                "ativo": cls._excel_bool(item.ativo),
            }
            for item in categorias
        ]

    @classmethod
    def _export_produtos(cls, tenant_id, escopo, empresa_id=None):
        produtos = ImportExportRepository.listar_produtos_empresa(
            tenant_id,
            AcessoEmpresaService.filtrar_empresa_ids(escopo),
            empresa_id=empresa_id,
        )
        return [
            {
                "empresa": item.empresa.nome_fantasia if item.empresa else "",
                "categoria": item.produto.categoria.nome if item.produto and item.produto.categoria else "",
                "nome": item.produto.nome if item.produto else "",
                "descricao": item.produto.descricao if item.produto else "",
                "codigo_barras": item.produto.codigo_barras if item.produto else "",
                "possui_ncm": cls._excel_bool(item.produto.possui_ncm if item.produto else False),
                "ncm": item.produto.ncm if item.produto else "",
                "estoque_atual": int(item.estoque_atual or 0),
                "estoque_minimo": int(item.estoque_minimo or 0),
                "valor_compra": cls._decimal_to_excel(item.valor_compra),
                "valor_venda": cls._decimal_to_excel(item.valor_venda),
                "data_validade": item.data_validade.isoformat() if item.data_validade else "",
                "ativo": cls._excel_bool(item.ativo),
            }
            for item in produtos
        ]

    @classmethod
    def _export_funcionarios(cls, tenant_id, escopo, empresa_id=None):
        vinculos = ImportExportRepository.listar_funcionarios_empresa(
            tenant_id,
            AcessoEmpresaService.filtrar_empresa_ids(escopo),
            empresa_id=empresa_id,
        )
        return [
            {
                "empresa": item.empresa.nome_fantasia if item.empresa else "",
                "role": item.funcionario.role.codigo if item.funcionario and item.funcionario.role else "",
                "nome": item.funcionario.nome if item.funcionario else "",
                "cpf": item.funcionario.cpf if item.funcionario else "",
                "usuario": item.funcionario.usuario if item.funcionario else "",
                "senha": "",
                "salario": cls._decimal_to_excel(item.funcionario.salario if item.funcionario else 0),
                "meta": cls._decimal_to_excel(item.funcionario.meta if item.funcionario else 0),
                "ativo": cls._excel_bool(item.funcionario.ativo if item.funcionario else False),
            }
            for item in vinculos
        ]

    @classmethod
    def _export_cupons(cls, tenant_id, escopo, empresa_id=None):
        del escopo, empresa_id
        return [
            {
                "nome": item.nome,
                "codigo": item.codigo,
                "data_validade": item.data_validade.isoformat(),
                "tipo_desconto": item.tipo_desconto.value,
                "valor_desconto": cls._decimal_to_excel(item.valor_desconto),
                "ativo": cls._excel_bool(item.ativo),
            }
            for item in ImportExportRepository.listar_cupons(tenant_id)
        ]

    @classmethod
    def _export_formas_pagamento(cls, tenant_id, escopo, empresa_id=None):
        del escopo, empresa_id
        return [
            {
                "nome": item.nome,
                "ativo": cls._excel_bool(item.ativo),
            }
            for item in ImportExportRepository.listar_formas_pagamento(tenant_id)
        ]

    @classmethod
    def _export_categorias_financeiras(cls, tenant_id, escopo, empresa_id=None):
        del escopo, empresa_id
        return [
            {
                "tipo_categoria": item.tipo_categoria.value,
                "nome": item.nome,
                "ativo": cls._excel_bool(item.ativo),
            }
            for item in ImportExportRepository.listar_categorias_financeiras(tenant_id)
        ]

    @classmethod
    def _fill_instructions_sheet(cls, sheet, config):
        sheet["A1"] = f"Modelo de importacao - {config['label']}"
        sheet["A1"].font = cls.TITLE_FONT
        sheet["A3"] = "Como usar"
        sheet["A3"].fill = cls.SECTION_FILL
        sheet["A3"].font = cls.HEADER_FONT

        instrucoes = [
            "1. Preencha somente a aba 'dados'.",
            "2. Nao altere o nome das colunas do cabecalho.",
            "3. Valores booleanos aceitos: SIM ou NAO.",
            "4. Datas devem seguir preferencialmente o formato YYYY-MM-DD.",
            "5. Quando um registro ja existir, o sistema atualiza os dados em vez de duplicar.",
        ]
        for offset, texto in enumerate(instrucoes, start=4):
            sheet[f"A{offset}"] = texto
            sheet[f"A{offset}"].alignment = Alignment(wrap_text=True)

        if config["label"] == "Produtos":
            sheet["A10"] = "Observacao"
            sheet["A10"].fill = cls.SECTION_FILL
            sheet["A10"].font = cls.HEADER_FONT
            sheet["A11"] = "Categorias novas informadas na planilha de produtos sao criadas automaticamente."
        elif config["label"] == "Funcionarios":
            sheet["A10"] = "Observacao"
            sheet["A10"].fill = cls.SECTION_FILL
            sheet["A10"].font = cls.HEADER_FONT
            sheet["A11"] = "A senha pode ficar em branco apenas para atualizar funcionarios ja existentes."

        sheet["D3"] = "Colunas"
        sheet["D3"].fill = cls.SECTION_FILL
        sheet["D3"].font = cls.HEADER_FONT
        sheet["D4"] = "campo"
        sheet["E4"] = "obrigatorio"
        sheet["F4"] = "exemplo"
        for cell in ("D4", "E4", "F4"):
            sheet[cell].fill = cls.HEADER_FILL
            sheet[cell].font = cls.HEADER_FONT

        for index, coluna in enumerate(config["columns"], start=5):
            sheet[f"D{index}"] = coluna["header"]
            sheet[f"E{index}"] = "SIM" if coluna["required"] else "NAO"
            sheet[f"F{index}"] = coluna.get("example") or ""

    @classmethod
    def _fill_reference_sheet(cls, sheet, entidade, tenant_id, escopo):
        row_cursor = 1
        blocks = cls._build_reference_blocks(entidade, tenant_id, escopo)

        if not blocks:
            sheet["A1"] = "Sem referencias adicionais para este modelo."
            return

        for title, headers, rows in blocks:
            sheet[f"A{row_cursor}"] = title
            sheet[f"A{row_cursor}"].fill = cls.SECTION_FILL
            sheet[f"A{row_cursor}"].font = cls.HEADER_FONT
            row_cursor += 1

            for idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_cursor, column=idx, value=header)
                cell.fill = cls.HEADER_FILL
                cell.font = cls.HEADER_FONT
            row_cursor += 1

            for values in rows:
                for idx, value in enumerate(values, start=1):
                    sheet.cell(row=row_cursor, column=idx, value=value)
                row_cursor += 1

            row_cursor += 2

    @classmethod
    def _build_reference_blocks(cls, entidade, tenant_id, escopo):
        empresa_ids = AcessoEmpresaService.filtrar_empresa_ids(escopo)
        empresas = ImportExportRepository.listar_empresas(tenant_id, empresa_ids)

        if entidade == "produtos":
            categorias = ImportExportRepository.listar_categorias_produto(tenant_id)
            return [
                (
                    "Empresas disponiveis",
                    ["nome_fantasia", "cnpj"],
                    [[item.nome_fantasia, item.cnpj] for item in empresas],
                ),
                (
                    "Categorias ja cadastradas",
                    ["nome", "descricao"],
                    [[item.nome, item.descricao or ""] for item in categorias],
                ),
            ]

        if entidade == "funcionarios":
            roles = ImportExportRepository.listar_roles(tenant_id)
            return [
                (
                    "Empresas disponiveis",
                    ["nome_fantasia", "cnpj"],
                    [[item.nome_fantasia, item.cnpj] for item in empresas],
                ),
                (
                    "Roles disponiveis",
                    ["codigo", "nome"],
                    [[item.codigo, item.nome] for item in roles],
                ),
            ]

        if entidade == "cupons":
            return [
                (
                    "Tipos de desconto aceitos",
                    ["tipo"],
                    [["PERCENTUAL"], ["VALOR"]],
                ),
            ]

        if entidade == "categorias_financeiras":
            return [
                (
                    "Tipos financeiros aceitos",
                    ["tipo_categoria"],
                    [["ENTRADA"], ["SAIDA"]],
                ),
            ]

        return [
            (
                "Empresas visiveis no seu acesso",
                ["nome_fantasia", "cnpj"],
                [[item.nome_fantasia, item.cnpj] for item in empresas],
            ),
        ]

    @classmethod
    def _get_data_sheet(cls, workbook):
        for sheet in workbook.worksheets:
            if sheet.title.strip().lower() == "dados":
                return sheet
        return workbook.active

    @classmethod
    def _build_header_map(cls, header_row, columns):
        actual_headers = [cls._normalize_header(value) for value in header_row]
        index_map = {}
        missing_required = []

        for coluna in columns:
            normalized = cls._normalize_header(coluna["header"])
            if normalized in actual_headers:
                index_map[coluna["key"]] = actual_headers.index(normalized)
            else:
                index_map[coluna["key"]] = None
                if coluna["required"]:
                    missing_required.append(coluna["header"])

        if missing_required:
            raise ValueError(
                "Colunas obrigatorias ausentes na planilha: " + ", ".join(missing_required)
            )

        return index_map

    @classmethod
    def _extract_row(cls, values, index_map):
        row = {}
        for key, index in index_map.items():
            row[key] = values[index] if index is not None and index < len(values) else None
        return row

    @classmethod
    def _row_is_empty(cls, values):
        return all(cls._optional_text(value) is None for value in values)

    @classmethod
    def _resolve_empresa(cls, tenant_id, escopo, identificador):
        empresa = ImportExportRepository.buscar_empresa_por_identificador(
            tenant_id,
            identificador,
            AcessoEmpresaService.filtrar_empresa_ids(escopo),
        )
        if not empresa:
            raise ValueError("Empresa nao encontrada ou fora do seu escopo de acesso.")
        AcessoEmpresaService.validar_empresa(empresa.id, escopo)
        return empresa

    @classmethod
    def _resolve_role(cls, tenant_id, identificador):
        role = ImportExportRepository.buscar_role_por_identificador(tenant_id, identificador)
        if not role:
            raise ValueError("Role nao encontrada.")
        return role

    @classmethod
    def _validate_operation(cls, entidade, escopo, operation):
        config = cls._get_entity_config(entidade)
        permission_codes = [
            cls.GENERIC_OPERATION_PERMISSIONS[operation],
            config[f"{operation}_permission"],
        ]
        if not all(AcessoEmpresaService.possui_permissao(escopo, codigo) for codigo in permission_codes):
            raise PermissionError("Voce nao possui permissao para executar esta operacao.")

    @classmethod
    def _pode_operar_entidade(cls, entidade, escopo, operation):
        config = cls._get_entity_config(entidade)
        permission_codes = [
            cls.GENERIC_OPERATION_PERMISSIONS[operation],
            config[f"{operation}_permission"],
        ]
        return all(AcessoEmpresaService.possui_permissao(escopo, codigo) for codigo in permission_codes)

    @classmethod
    def _get_entity_config(cls, entidade):
        if entidade not in cls.ENTITY_DEFINITIONS:
            raise ValueError("Entidade de importacao/exportacao invalida.")
        return cls.ENTITY_DEFINITIONS[entidade]

    @staticmethod
    def _normalize_header(value):
        raw = str(value or "").strip().lower().replace(" ", "_")
        return unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")

    @staticmethod
    def _required_text(value, field_name):
        texto = ImportExportService._optional_text(value)
        if not texto:
            raise ValueError(f"Campo obrigatorio ausente: {field_name}.")
        return texto

    @staticmethod
    def _optional_text(value):
        if value is None:
            return None
        texto = str(value).strip()
        return texto or None

    @staticmethod
    def _parse_bool(value, default=True):
        if value in (None, ""):
            return default
        if isinstance(value, bool):
            return value

        normalized = str(value).strip().lower()
        if normalized in {"1", "true", "sim", "s", "yes", "y", "ativo"}:
            return True
        if normalized in {"0", "false", "nao", "não", "n", "no", "inativo"}:
            return False

        raise ValueError(f"Valor booleano invalido: {value}. Use SIM ou NAO.")

    @staticmethod
    def _parse_int(value, default=0, field_name="valor"):
        if value in (None, ""):
            return default

        if isinstance(value, int):
            parsed = value
        elif isinstance(value, float):
            parsed = int(value)
        else:
            texto = str(value).strip().replace(".", "").replace(",", "")
            try:
                parsed = int(texto)
            except ValueError:
                raise ValueError(f"Valor invalido para {field_name}.")

        if parsed < 0:
            raise ValueError(f"{field_name} nao pode ser negativo.")
        return parsed

    @staticmethod
    def _parse_decimal(value, default="0.00", field_name="valor"):
        if value in (None, ""):
            if default is None:
                raise ValueError(f"Campo obrigatorio ausente: {field_name}.")
            value = default

        if isinstance(value, Decimal):
            decimal_value = value
        elif isinstance(value, (int, float)):
            decimal_value = Decimal(str(value))
        else:
            texto = str(value).strip()
            if "," in texto and "." in texto:
                if texto.rfind(",") > texto.rfind("."):
                    texto = texto.replace(".", "").replace(",", ".")
                else:
                    texto = texto.replace(",", "")
            elif "," in texto:
                texto = texto.replace(".", "").replace(",", ".")
            try:
                decimal_value = Decimal(texto)
            except (InvalidOperation, ValueError):
                raise ValueError(f"Valor invalido para {field_name}.")

        if decimal_value < 0:
            raise ValueError(f"{field_name} nao pode ser negativo.")

        return decimal_value.quantize(Decimal("0.01"))

    @staticmethod
    def _parse_optional_date(value, field_name):
        if value in (None, ""):
            return None
        return ImportExportService._parse_required_date(value, field_name)

    @staticmethod
    def _parse_required_date(value, field_name):
        if value in (None, ""):
            raise ValueError(f"Campo obrigatorio ausente: {field_name}.")

        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        texto = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue

        raise ValueError(f"Data invalida para {field_name}. Use YYYY-MM-DD.")

    @staticmethod
    def _parse_tipo_desconto(value):
        texto = str(value or "").strip().upper()
        try:
            return TipoDesconto[texto]
        except KeyError:
            raise ValueError("tipo_desconto invalido. Use PERCENTUAL ou VALOR.")

    @staticmethod
    def _parse_tipo_categoria(value):
        texto = str(value or "").strip().upper()
        try:
            return TipoCategoriaFinanceira[texto]
        except KeyError:
            raise ValueError("tipo_categoria invalido. Use ENTRADA ou SAIDA.")

    @staticmethod
    def _excel_bool(value):
        return "SIM" if bool(value) else "NAO"

    @staticmethod
    def _decimal_to_excel(value):
        try:
            return float(Decimal(str(value or 0)).quantize(Decimal("0.01")))
        except (InvalidOperation, ValueError):
            return 0.0

    @classmethod
    def _style_header_row(cls, sheet, row_number):
        for cell in sheet[row_number]:
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_fit_sheet(sheet):
        dimensions = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                dimensions[cell.column_letter] = max(dimensions.get(cell.column_letter, 0), min(length + 4, 45))

        for column_letter, width in dimensions.items():
            sheet.column_dimensions[column_letter].width = width

    @staticmethod
    def _workbook_to_bytes(workbook):
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
