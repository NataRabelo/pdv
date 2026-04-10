import os
import unittest
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app import create_app
from app.extensions import db
from app.models.db import (
    CategoriaProduto,
    Empresa,
    Funcionario,
    FuncionarioEmpresa,
    Tenant,
    TipoEmpresa,
)
from app.security.password import hash_password
from app.services.acesso_empresa_service import AcessoEmpresaService
from app.services.import_export_service import ImportExportService
from app.services.tenant_bootstrap_service import TenantBootstrapService


class ImportExportServiceTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ["DATABASE_URL"] = "sqlite:///test_import_export.db"
        cls.app = create_app()
        cls.app_context = cls.app.app_context()
        cls.app_context.push()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.drop_all()
        cls.app_context.pop()

    def setUp(self):
        db.drop_all()
        db.create_all()

        tenant = Tenant(nome="Tenant Importacao")
        db.session.add(tenant)
        db.session.commit()

        roles = TenantBootstrapService.garantir_permissoes_e_roles(tenant.id)
        TenantBootstrapService.garantir_cadastros_operacionais(tenant.id)

        empresa = Empresa(
            tenant_id=tenant.id,
            cnpj="98.765.432/0001-99",
            razao_social="BlueOcean Importacoes LTDA",
            nome_fantasia="BlueOcean Centro",
            tipo_empresa=TipoEmpresa.MATRIZ,
            ativo=True,
        )
        db.session.add(empresa)
        db.session.flush()

        funcionario = Funcionario(
            tenant_id=tenant.id,
            role_id=roles["administrador"].id,
            nome="Admin Importacao",
            cpf="555.444.333-22",
            usuario="admin.importacao",
            senha_hash=hash_password("123456"),
            ativo=True,
        )
        db.session.add(funcionario)
        db.session.flush()

        db.session.add(
            FuncionarioEmpresa(
                tenant_id=tenant.id,
                funcionario_id=funcionario.id,
                empresa_id=empresa.id,
                ativo=True,
            )
        )

        categoria = CategoriaProduto(
            tenant_id=tenant.id,
            nome="Padaria",
            descricao="Itens de padaria",
            ativo=True,
        )
        db.session.add(categoria)
        db.session.commit()

        self.tenant = tenant
        self.empresa = empresa
        self.funcionario = funcionario
        self.escopo = AcessoEmpresaService.obter_escopo(funcionario.id, tenant.id)

    def test_template_produtos_contem_abas_e_headers_esperados(self):
        arquivo = ImportExportService.gerar_template("produtos", self.tenant.id, self.escopo)
        workbook = load_workbook(BytesIO(arquivo["content"]))

        self.assertIn("dados", workbook.sheetnames)
        self.assertIn("instrucoes", workbook.sheetnames)
        self.assertIn("referencias", workbook.sheetnames)

        headers = [cell.value for cell in workbook["dados"][1]]
        self.assertEqual(headers[:5], ["empresa", "categoria", "nome", "descricao", "codigo_barras"])

    def test_importacao_de_produto_cria_categoria_e_vinculo_por_empresa(self):
        resultado = self._importar_produto_teste()

        categoria = CategoriaProduto.query.filter_by(
            tenant_id=self.tenant.id,
            nome="Bebidas geladas",
        ).first()
        self.assertIsNotNone(categoria)
        self.assertEqual(resultado["sucesso"], 1)
        self.assertEqual(resultado["falhas"], 0)

    def test_exportacao_de_produtos_retorna_linha_com_empresa_e_valores(self):
        self._importar_produto_teste()

        arquivo = ImportExportService.exportar_entidade(
            "produtos",
            self.tenant.id,
            self.escopo,
            empresa_id=self.empresa.id,
        )
        workbook = load_workbook(BytesIO(arquivo["content"]), data_only=True)
        dados = list(workbook["dados"].iter_rows(values_only=True))

        self.assertGreaterEqual(len(dados), 2)
        self.assertEqual(dados[1][0], self.empresa.nome_fantasia)
        self.assertEqual(dados[1][2], "Suco de Uva 1L")
        self.assertEqual(dados[1][7], 18)

    def _importar_produto_teste(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "dados"
        sheet.append([
            "empresa",
            "categoria",
            "nome",
            "descricao",
            "codigo_barras",
            "possui_ncm",
            "ncm",
            "estoque_atual",
            "estoque_minimo",
            "valor_compra",
            "valor_venda",
            "data_validade",
            "ativo",
        ])
        sheet.append([
            self.empresa.nome_fantasia,
            "Bebidas geladas",
            "Suco de Uva 1L",
            "Importado pelo teste",
            "7891234567890",
            "NAO",
            "",
            18,
            4,
            "5,90",
            "9,90",
            "2026-12-31",
            "SIM",
        ])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.filename = "produtos.xlsx"

        resultado = ImportExportService.importar_entidade(
            "produtos",
            buffer,
            self.tenant.id,
            self.escopo,
            self.funcionario.id,
        )
        return resultado


if __name__ == "__main__":
    unittest.main()
